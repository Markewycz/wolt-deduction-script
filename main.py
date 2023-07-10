# -*- coding: utf-8 -*
import cv2
from pytesseract import pytesseract
from PIL import Image
from PIL.ExifTags import TAGS
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os

class ProcessImage:

  def image_to_text(img_path):
    img = cv2.imread(img_path)
    img = img[1610:1870,95:890,:]

    # Get the base directory relative to the script
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # Construct the absolute path to the Tesseract-OCR executable
    tesseract_path = os.path.join(base_dir, "Tesseract-OCR", "tesseract.exe")

    
    pytesseract.tesseract_cmd = tesseract_path
    txt = pytesseract.image_to_string(img, lang="dan")
    lines = txt.split("\n")

    address1, address2, blank, total_km = lines[:4]
    total_formatted = total_km.split(" ")

    total_route = None # Initialize with a default value

  # Checks if total_formatted has length of 5 or 6
    if len(total_formatted) == 6:
      splitted = total_formatted[1].split("+")
      if len(splitted) > 1:
        total_route = splitted[1]
    elif len(total_formatted) == 5:
      total_route = total_formatted[3]
    else:
       total_route = None

    return address1, address2, total_route
  

class Metadata:
  def get_date(img_path):
    # Get image
    img_data = Image.open(img_path)

    # Get EXIF data from image
    exif_data = img_data._getexif()

    # Check if the EXIF data exists and contains the creation date
    if exif_data is not None:
        for tag_id, value in exif_data.items():
            tag_name = TAGS.get(tag_id, tag_id)
            if tag_name == 'DateTimeOriginal':
                creation_array = value.split(" ")
                creation_date = creation_array[0]
                break
    else:
        print("No EXIF data found.")
    return creation_date

class ImportData:
  temp_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "images")
  
  @staticmethod
  def sort_spreadsheet():
    # Load the workbook
    workbook = openpyxl.load_workbook("data.xlsx")
    sheet = workbook.active

    # Determine the column index for the date column
    date_column = 'A'
    date_column_index = column_index_from_string(date_column)

    # Sort the data based on the date column starting from the second row
    sorted_rows = sorted(sheet.iter_rows(min_row=2, min_col=date_column_index, values_only=True), key=lambda x: x[0])

    # Clear the existing data in the spreadsheet from the second row
    sheet.delete_rows(2, sheet.max_row)

    # Write the sorted data back to the spreadsheet starting from the second row
    for row in sorted_rows:
        sheet.append(row)

    # Calculate the sum of values in column D starting from the second cell
    sum_formula = f"SUM(D2:D{sheet.max_row})"
    sheet["E2"] = f"={sum_formula}"

    # Save the sorted spreadsheet
    workbook.save("data.xlsx")

  def modify_spreadsheet(address1, address2, total, date):
    # Open the existing spreadsheet or create a new one
    excel_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")
    if not os.path.exists(excel_file_path):
      workbook = openpyxl.Workbook()
      workbook.save(excel_file_path)

    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Get the next empty row in the spreadsheet
    next_row = sheet.max_row + 1

    # Write the values to the corresponding cells
    sheet.cell(row=next_row, column=1).value = date
    sheet.cell(row=next_row, column=2).value = address1
    sheet.cell(row=next_row, column=3).value = address2
    sheet.cell(row=next_row, column=4).value = float(total)

    # Save the spreadsheet
    workbook.save('data.xlsx')

    # Sort the spreadsheet
    ImportData.sort_spreadsheet()

  def process_images():
    # Get the list of image files in the temporary folder
    image_files = [file for file in os.listdir(ImportData.temp_folder) if file.endswith(".jpg") or file.endswith(".PNG")]

    # Process each image file
    for filename in image_files:
        # Construct the full path of the image file
        img_path = os.path.join(ImportData.temp_folder, filename)

        # Process data from image
        a, b, total = ProcessImage.image_to_text(img_path)
        date = Metadata.get_date(img_path)

        # Export to existing excel file
        ImportData.modify_spreadsheet(a, b, total, date)

if __name__ == "__main__":
    os.getcwd()
    ImportData.process_images()
